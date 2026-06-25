"""
PERFILES DE CALIDAD Y CAUDAL - RIO RISARALDA
UAO + ASOCANA | Proyecto Corredor Biologico 890K

Genera:
  1. Perfil longitudinal de calidad del agua (un PNG por parametro)
  2. Graficas de caudal (2 estaciones: EHT-Rio Risaralda y Casa Maquinas)
     - Caudal promedio mensual (barras)
     - Caudal mensual 2025 vs 2026 (lineas por ano)
     - Curva de duracion de caudales

Salida: C:\\Users\\ASUS\\Desktop\\Work\\Proyecto - Corredor Biologico\\
        Fase I\\Bases de datos\\Estaciones\\Tributarios\\RIO RISARALDA\\

Requiere: pip install pandas openpyxl matplotlib numpy
"""

import sys
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from pathlib import Path
import warnings
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook

warnings.filterwarnings("ignore")

# ── Configuracion ──────────────────────────────────────────────
BASE_DIR  = Path(r"C:\Users\ASUS\Desktop\Work\Proyecto - Corredor Biológico\Fase I\Bases de datos\Estaciones\Tributarios\RIO RISARALDA")
EXCEL_CAL = BASE_DIR / "PUNTOS_DE_MONITOREO.xlsx"
EXCEL_RR  = BASE_DIR / "CAUDAL" / "Rio_Risaralda_caudal.xlsx"
EXCEL_CM  = BASE_DIR / "CAUDAL" / "Casa_Maquinas_Rda_caudal.xlsx"
DPI       = 180

MESES_ABR = ["ENE","FEB","MAR","ABR","MAY","JUN","JUL","AGO","SEP","OCT","NOV","DIC"]

# ── Estaciones de calidad ──────────────────────────────────────
# km_vis: posicion visual en el eje X (separada para evitar superposicion)
# km_real: distancia real desde SUP-050 (se muestra en la etiqueta)
# label: texto que aparece bajo el punto en el eje X
ESTACIONES_CAL = {
    "SUP-050": {"km_vis":  0.00, "km_real":  0.00, "label": "SUP-050\n(0 km)",              "hoja": "SUP-050"},
    "SUP-195": {"km_vis":  6.00, "km_real":  1.94, "label": "SUP-195\n(1.94 km)",           "hoja": "SUP-195."},
    "SUP-049": {"km_vis": 14.00, "km_real": 11.88, "label": "SUP-049\nPte. Umbria\n(11.88 km)", "hoja": "SUP-049"},
    "SUP-123": {"km_vis": 22.00, "km_real": 17.08, "label": "SUP-123\nR. Guatica\n(17.08 km)",  "hoja": "SUP-123"},
    "SUP-048": {"km_vis": 38.00, "km_real": 36.33, "label": "SUP-048\n(36.33 km)",          "hoja": "SUP-048"},
    "SUP-047": {"km_vis": 47.00, "km_real": 36.60, "label": "SUP-047\n(36.60 km)",          "hoja": "SUP-047"},
    "SUP-058": {"km_vis": 62.00, "km_real": 59.07, "label": "SUP-058\nPte. Negro\n(59.07 km)", "hoja": "SUP-058"},
    "SUP-241": {"km_vis": 79.35, "km_real": 79.35, "label": "SUP-241\nDesemboc.\n(79.35 km)","hoja": "SUP-241"},
    "SUP-105": {"km_vis": 88.00, "km_real": 79.60, "label": "SUP-105\nR. Cauca\n(79.60 km)", "hoja": "SUP-105."},
}

# Estaciones de caudal
ESTACIONES_CAU = {
    "Rio Risaralda (EHT)": {"km": 15.80, "excel": EXCEL_RR, "color": "#1565C0"},
    "Casa Maquinas":        {"km": 14.89, "excel": EXCEL_CM, "color": "#6A1B9A"},
}

# Parametros de calidad
PARAMETROS = {
    "OD":        {"col": "Oxigeno Disuelto (mg O2/L)",                               "label": "OD (mg O2/L)",           "color": "#1565C0", "umbral": 4.0,  "umbral_label": "Decreto 1076 (4 mg/L)"},
    "DBO5":      {"col": "Demanda Bioquimica de Oxigeno (total o soluble) (mg O2/L)", "label": "DBO5 (mg O2/L)",         "color": "#C62828", "umbral": None, "umbral_label": None},
    "SST":       {"col": "Sólidos Suspendidos Totales (mg/L)",                        "label": "SST (mg/L)",             "color": "#6A1B9A", "umbral": None, "umbral_label": None},
    "Turbiedad": {"col": "Turbiedad (NTU)",                                           "label": "Turbiedad (NTU)",        "color": "#EF6C00", "umbral": None, "umbral_label": None},
    "NT":        {"col": "Nitrogeno Total (mgN/L)",                                   "label": "N Total (mgN/L)",        "color": "#2E7D32", "umbral": None, "umbral_label": None},
    "PT":        {"col": "Fosforo Total (mgP-PO4-3/L)",                               "label": "P Total (mgP/L)",        "color": "#AD1457", "umbral": 0.1,  "umbral_label": "Ref. eutrofizacion (0.1 mgP/L)"},
    "pH":        {"col": "pH (Uds. pH) (Uds. pH)",                                    "label": "pH",                    "color": "#00695C", "umbral": None, "umbral_label": None},
    "Cond":      {"col": "Conductividad (uS/cm) (uS/cm)",                             "label": "Conductividad (uS/cm)", "color": "#4E342E", "umbral": None, "umbral_label": None},
    "Colif":     {"col": "Coliformes Totales (E + 1) (NMP/100mL)",                    "label": "Col. Totales (NMP/100mL)","color": "#37474F","umbral": None, "umbral_label": None},
}

# ── Funciones auxiliares ───────────────────────────────────────
def limpiar_valor(v):
    if pd.isna(v): return np.nan
    s = str(v).strip()
    if s in ("-", "", "nd", "ND"): return np.nan
    if s.startswith("<"):
        try: return float(s[1:].strip()) 
        except: return np.nan
    try: return float(s.replace(",", "."))
    except: return np.nan


def leer_caudal_excel(excel_path: Path) -> pd.DataFrame:
    """Lee el Excel de caudal (estructura: fila2=meses, fila3=CMax/CMed/CMin)."""
    wb   = load_workbook(excel_path, data_only=True)
    rows = []
    for sheet_name in wb.sheetnames:
        try: ano = int(sheet_name)
        except: continue
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        # Fila 2 (indice 1): meses — cada 3 columnas
        fila_mes = all_rows[1]
        meses_idx = {}
        for ci, val in enumerate(fila_mes):
            if val in MESES_ABR:
                meses_idx[ci]   = (val, "CMax")
                meses_idx[ci+1] = (val, "CMed")
                meses_idx[ci+2] = (val, "CMin")
        # Filas de datos (a partir de fila 4, indice 3)
        for row in all_rows[3:]:
            if row[0] is None: continue
            try: dia = int(row[0])
            except: continue
            if not 1 <= dia <= 31: continue
            for ci, (mes, tipo) in meses_idx.items():
                if ci < len(row) and isinstance(row[ci], (int, float)):
                    rows.append({"ano": ano, "mes": mes, "dia": dia,
                                 "tipo": tipo, "valor": float(row[ci])})
    return pd.DataFrame(rows)


def estilo_base(fig, ax, color="#1565C0"):
    fig.patch.set_facecolor("#F7F9FC")
    ax.set_facecolor("#F7F9FC")
    ax.grid(axis="y", color="#E0E0E0", linewidth=0.7, linestyle="--", zorder=0)
    for spine in ax.spines.values():
        spine.set_visible(False)
    ax.spines["bottom"].set_visible(True)
    ax.spines["bottom"].set_color("#CCCCCC")
    ax.spines["left"].set_visible(True)
    ax.spines["left"].set_color(color)
    ax.spines["left"].set_linewidth(3)


# ══════════════════════════════════════════════════════════════════
# BLOQUE 1: PERFILES LONGITUDINALES DE CALIDAD
# ══════════════════════════════════════════════════════════════════
print("=" * 60)
print("BLOQUE 1: PERFILES DE CALIDAD - RIO RISARALDA")
print("=" * 60)

xl = pd.ExcelFile(EXCEL_CAL)

# Leer y normalizar columnas de cada hoja
def leer_hoja(xl, hoja):
    df = pd.read_excel(xl, sheet_name=hoja)
    # Normalizar nombres de columna: quitar espacios, tildes comunes
    nuevas = []
    for c in df.columns:
        c2 = c.strip()
        c2 = c2.replace("Oxígeno","Oxigeno").replace("Bioquímica","Bioquimica")
        c2 = c2.replace("Nitrógeno","Nitrogeno").replace("Fósforo","Fosforo")
        c2 = c2.replace("Conductividad eléctrica","Conductividad (uS/cm)")
        nuevas.append(c2)
    df.columns = nuevas
    return df

# Calcular promedios por estacion
resumen = []
for cod, info in ESTACIONES_CAL.items():
    df = leer_hoja(xl, info["hoja"])
    fila = {"codigo": cod, "km_vis": info["km_vis"],
            "km_real": info["km_real"], "label": info["label"]}
    for key, pinfo in PARAMETROS.items():
        col = pinfo["col"]
        # Buscar columna con coincidencia parcial si no hay exacta
        col_found = None
        if col in df.columns:
            col_found = col
        else:
            for c in df.columns:
                if col.lower()[:15] in c.lower():
                    col_found = c
                    break
        if col_found:
            vals = df[col_found].apply(limpiar_valor).dropna()
            if key == "DBO5":
                vals = vals[vals <= 50]
            fila[key] = vals.mean() if len(vals) > 0 else np.nan
        else:
            fila[key] = np.nan
    resumen.append(fila)

dfr = pd.DataFrame(resumen).sort_values("km_vis").reset_index(drop=True)
km_vis    = dfr["km_vis"].values
etiquetas = dfr["label"].values
x_lim_max = km_vis[-1] + 5

# Generar un PNG por parametro
for key, pinfo in PARAMETROS.items():
    y     = dfr[key].values
    mask  = ~np.isnan(y)
    color = pinfo["color"]

    fig, ax = plt.subplots(figsize=(15, 6))
    estilo_base(fig, ax, color)

    if mask.sum() >= 2:
        ax.fill_between(km_vis[mask], y[mask], alpha=0.07, color=color)
        ax.plot(km_vis[mask], y[mask], "-o", color=color, linewidth=2.2,
                markersize=8, markerfacecolor="white", markeredgecolor=color,
                markeredgewidth=2.2, zorder=4)

    # Puntos sin dato
    for km, val in zip(km_vis, y):
        if np.isnan(val):
            ax.axvline(km, color="#BDBDBD", linewidth=0.8, linestyle=":", alpha=0.6)

    # Umbral normativo
    if pinfo["umbral"] is not None:
        ax.axhline(pinfo["umbral"], color="#F44336", linewidth=1.5,
                   linestyle="--", alpha=0.85, zorder=3,
                   label=pinfo["umbral_label"])
        ax.legend(fontsize=8, framealpha=0.8, loc="upper right")

    # Anotaciones de valor
    for km, val in zip(km_vis[mask], y[mask]):
        ax.annotate(f"{val:.2f}", (km, val),
                    textcoords="offset points", xytext=(0, 9),
                    ha="center", fontsize=8, color=color, fontweight="bold")

    # Lineas verticales de referencia
    for km in km_vis:
        ax.axvline(km, color="#E8EAF6", linewidth=0.8, zorder=1)

    ax.set_xlim(-3, x_lim_max)
    ax.set_xticks(km_vis)
    ax.set_xticklabels(etiquetas, fontsize=8.5, ha="center", linespacing=1.4)
    ax.set_xlabel("Distancia aproximada desde SUP-050 (km)  |  Aguas arriba -> Aguas abajo",
                  fontsize=10, labelpad=12)
    ax.set_ylabel(pinfo["label"], fontsize=11, color=color, fontweight="bold")
    ax.tick_params(axis="y", labelsize=9)
    ax.set_title(
        f"Perfil Longitudinal - {pinfo['label']}\n"
        f"Rio Risaralda  |  Promedio historico por estacion  |  Fuente: CARDER",
        fontsize=12, fontweight="bold", color="#1A237E", pad=12
    )

    out = BASE_DIR / f"perfil_calidad_{key}.png"
    plt.tight_layout()
    plt.savefig(out, dpi=DPI, bbox_inches="tight", facecolor="#F7F9FC")
    plt.close()
    print(f"  [OK] perfil_calidad_{key}.png")

print(f"  {len(PARAMETROS)} perfiles guardados\n")


# ══════════════════════════════════════════════════════════════════
# BLOQUE 2: GRAFICAS DE CAUDAL
# ══════════════════════════════════════════════════════════════════
print("=" * 60)
print("BLOQUE 2: GRAFICAS DE CAUDAL")
print("=" * 60)

for est_nombre, est_info in ESTACIONES_CAU.items():
    excel_path = est_info["excel"]
    color      = est_info["color"]
    print(f"\n  Procesando: {est_nombre}")

    df_cau = leer_caudal_excel(excel_path)
    if df_cau.empty:
        print(f"  [AVISO] Sin datos en {excel_path.name}")
        continue

    df_med = df_cau[df_cau["tipo"] == "CMed"].copy()
    df_med["mes_num"] = df_med["mes"].map({m: i+1 for i, m in enumerate(MESES_ABR)})

    nombre_archivo = (est_nombre.replace(" ", "_")
                                .replace("(", "").replace(")", "")
                                .replace("/", "_"))

    # ── Grafica 1: Caudal promedio mensual ───────────────────────
    prom_mes = (df_med.groupby(["mes_num","mes"])["valor"]
                .mean().reset_index().sort_values("mes_num"))

    fig, ax = plt.subplots(figsize=(13, 6))
    estilo_base(fig, ax, color)

    norm    = plt.Normalize(prom_mes["valor"].min(), prom_mes["valor"].max())
    colores = plt.cm.YlGnBu(norm(prom_mes["valor"].values))
    bars    = ax.bar(prom_mes["mes"], prom_mes["valor"], color=colores,
                     edgecolor="white", linewidth=0.7, width=0.65, zorder=3)

    for bar, val in zip(bars, prom_mes["valor"]):
        ax.text(bar.get_x() + bar.get_width()/2,
                bar.get_height() + prom_mes["valor"].max()*0.015,
                f"{val:.2f}", ha="center", va="bottom",
                fontsize=8.5, fontweight="bold", color="#333333")

    prom_total = prom_mes["valor"].mean()
    ax.axhline(prom_total, color="#E63946", linewidth=1.4, linestyle="--", zorder=4,
               label=f"Promedio periodo: {prom_total:.2f} m3/s")
    ax.legend(fontsize=9, framealpha=0.8)
    ax.set_xlabel("Mes", fontsize=10, labelpad=8)
    ax.set_ylabel("Caudal Medio (m3/s)", fontsize=11, fontweight="bold")
    ax.set_title(f"Caudal Promedio Mensual - {est_nombre}\n"
                 f"Rio Risaralda  |  2025-2026  |  Fuente: CARDER / REDH",
                 fontsize=12, fontweight="bold", color="#1A237E", pad=12)
    ax.tick_params(axis="both", labelsize=9)

    out = BASE_DIR / f"caudal_prom_mensual_{nombre_archivo}.png"
    plt.tight_layout()
    plt.savefig(out, dpi=DPI, bbox_inches="tight", facecolor="#F7F9FC")
    plt.close()
    print(f"    [OK] caudal_prom_mensual_{nombre_archivo}.png")

    # ── Grafica 2: Caudal mensual 2025 vs 2026 ───────────────────
    fig, ax = plt.subplots(figsize=(13, 6))
    estilo_base(fig, ax, color)

    colores_ano = {2025: "#1565C0", 2026: "#E65100"}
    marcadores  = {2025: "o",       2026: "s"}

    for ano in sorted(df_med["ano"].unique()):
        df_ano = (df_med[df_med["ano"] == ano]
                .groupby(["mes_num", "mes"])["valor"]
                .mean().reset_index()
                .sort_values("mes_num"))  # ordenar por numero de mes, no por nombre
        if df_ano.empty:
            continue
        # Usar mes_num como eje X para garantizar orden correcto
        ax.plot(df_ano["mes_num"], df_ano["valor"],
                "-" + marcadores.get(ano, "o"),
                color=colores_ano.get(ano, "#333333"),
                linewidth=2, markersize=7, markerfacecolor="white",
                markeredgewidth=2, label=str(ano), zorder=4)

    # Poner etiquetas de mes en el eje X
    ax.set_xticks(range(1, 13))
    ax.set_xticklabels(MESES_ABR, fontsize=9)
    ax.set_xlim(0.5, 12.5)
    ax.legend(fontsize=10, framealpha=0.85, title="Ano")
    ax.set_xlabel("Mes", fontsize=10, labelpad=8)
    ax.set_ylabel("Caudal Medio (m3/s)", fontsize=11, fontweight="bold")
    ax.set_title(f"Variacion Mensual de Caudal - {est_nombre}\n"
                f"Rio Risaralda  |  2025 vs 2026  |  Fuente: CARDER / REDH",
                fontsize=12, fontweight="bold", color="#1A237E", pad=12)
    ax.tick_params(axis="both", labelsize=9)
    # ── Grafica 3: Curva de duracion de caudales ─────────────────
    todos = df_med["valor"].dropna().values
    if len(todos) < 30:
        print(f"    [AVISO] Pocos datos para curva de duracion ({len(todos)} valores)")
        continue

    datos_ord   = np.sort(todos)[::-1]
    n           = len(datos_ord)
    permanencia = np.arange(1, n+1) / n * 100
    u_verano    = float(np.interp(70, permanencia, datos_ord))
    u_invierno  = float(np.interp(30, permanencia, datos_ord))

    fig, ax = plt.subplots(figsize=(12, 6))
    estilo_base(fig, ax, color)

    ax.axvspan(0,  30, alpha=0.07, color="#2563EB")
    ax.axvspan(30, 70, alpha=0.07, color="#16A34A")
    ax.axvspan(70, 100, alpha=0.07, color="#F59E0B")

    ax.plot(permanencia, datos_ord, color="#1E3A5F", linewidth=1.8, zorder=4)
    ax.axvline(30, color="#2563EB", linewidth=1.4, linestyle="--", zorder=5)
    ax.axvline(70, color="#F59E0B", linewidth=1.4, linestyle="--", zorder=5)
    ax.axhline(u_invierno, color="#2563EB", linewidth=1.0, linestyle=":", alpha=0.7, zorder=5)
    ax.axhline(u_verano,   color="#F59E0B", linewidth=1.0, linestyle=":", alpha=0.7, zorder=5)

    ymax = datos_ord.max()
    ax.text(15,  ymax*0.95, "Invierno",   ha="center", fontsize=9, color="#2563EB", fontweight="bold")
    ax.text(50,  ymax*0.95, "Transicion", ha="center", fontsize=9, color="#16A34A", fontweight="bold")
    ax.text(85,  ymax*0.95, "Verano",     ha="center", fontsize=9, color="#F59E0B", fontweight="bold")
    ax.text(71, u_invierno, f"Q30={u_invierno:.2f} m3/s", fontsize=8, color="#2563EB", va="bottom")
    ax.text(71, u_verano,   f"Q70={u_verano:.2f} m3/s",   fontsize=8, color="#F59E0B", va="bottom")

    ax.set_xlabel("Porcentaje de permanencia (%)", fontsize=10, labelpad=8)
    ax.set_ylabel("Caudal Medio (m3/s)", fontsize=11, fontweight="bold")
    ax.set_title(f"Curva de Duracion de Caudales - {est_nombre}\n"
                 f"Rio Risaralda  |  2025-2026  |  Fuente: CARDER / REDH",
                 fontsize=12, fontweight="bold", color="#1A237E", pad=12)
    ax.set_xlim(0, 100)
    ax.xaxis.set_major_formatter(mticker.FormatStrFormatter("%g%%"))
    ax.tick_params(axis="both", labelsize=9)

    out = BASE_DIR / f"curva_duracion_{nombre_archivo}.png"
    plt.tight_layout()
    plt.savefig(out, dpi=DPI, bbox_inches="tight", facecolor="#F7F9FC")
    plt.close()
    print(f"    [OK] curva_duracion_{nombre_archivo}.png")

    # Guardar umbrales
    txt = BASE_DIR / f"umbrales_caudal_{nombre_archivo}.txt"
    txt.write_text(
        f"Estacion: {est_nombre}\n"
        f"Datos: {len(todos)} valores diarios\n"
        f"Periodo: 2025-2026\n\n"
        f"Umbral Invierno (Q30%): {u_invierno:.2f} m3/s\n"
        f"Umbral Verano   (Q70%): {u_verano:.2f} m3/s\n",
        encoding="utf-8"
    )
    print(f"    [OK] umbrales_caudal_{nombre_archivo}.txt")

print("\n" + "=" * 60)
print("[LISTO] Todas las figuras generadas.")
print(f"Carpeta: {BASE_DIR}")
print("=" * 60)