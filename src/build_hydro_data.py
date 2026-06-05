# -*- coding: utf-8 -*-
"""build_hydro_data.py — Preparación de datos de estaciones hidrométricas.

Se ejecuta UNA SOLA VEZ desde la terminal para generar los archivos estáticos
que el visor web consume:

    python src/build_hydro_data.py

Lee la carpeta fuente externa (Excel de caudal diario + umbrales + CDC por
estación), la cruza con la tabla de metadatos embebida abajo, y produce:

    data/hydrology/estaciones_hidro.json          (metadatos + estadísticas)
    data/hydrology/<NOMBRE>/caudal_diario.csv      (serie diaria por estación)
    data/hydrology/<NOMBRE>/curva_duracion_caudales.png  (si existe)
"""

import os
import re
import sys
import csv
import json
import shutil
import datetime

import openpyxl

# La consola de Windows usa cp1252 por defecto; forzar UTF-8 para los prints.
try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, ValueError):
    pass

# ── Rutas ──────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
SRC_DIR = os.path.join(
    PROJECT_DIR, 'data', 'databases', 'hidrology',
    'Estaciones Hidroclimatológicas - Río Cauca',
)
OUT_DIR = os.path.join(PROJECT_DIR, 'data', 'hydrology')

# Estaciones sin datos reales de caudal — excluir completamente
EXCLUIR = {'LA PRIMAVERA', 'LA FLORESTA'}

# Nombre de visualización limpio para cada estación
NOMBRE_DISPLAY = {
    'ANACARO':                   'Anacaro',
    'GARZONERO NORTE':           'Garzonero Norte',
    'GARZONERO SUR (S - 2023)':  'Garzonero Sur',
    'GUAYABAL (HASTA 2022)':     'Guayabal',
    'HORMIGUERO':                'Hormiguero',
    'JUAN DIAZ':                 'Juan Díaz',
    'JUANCHITO (S HASTA 2017)':  'Juanchito',
    'LA BALSA':                  'La Balsa',
    'LA BOLSA':                  'La Bolsa',
    'LA VICTORIA':               'La Victoria',
    'MEDICANOA':                 'Mediacanoa',
    'PAN DE AZUCAR':             'Pan de Azúcar',
    'PASO LA TORRE':             'Paso La Torre',
    'PUERTO MALLARINO':          'Puerto Mallarino',
    'PUERTO PEDRERO (S - 2017)': 'Puerto Pedrero',
    'TABLANCA':                  'Tablanca',
}

# ── Tabla de metadatos por estación (nombre de carpeta → atributos) ────
META = {
    'PAN DE AZUCAR':            dict(latitud=2.736424, longitud=-76.713214, tipo='Limnigráfica',  estado='Activa',     tiene_calidad=False, estacion_calidad=None),
    'TABLANCA':                 dict(latitud=3.127533, longitud=-76.576532, tipo='Limnimétrica',  estado='Suspendida', tiene_calidad=False, estacion_calidad=None),
    'LA BALSA':                 dict(latitud=3.09235,  longitud=-76.599022, tipo='Limnigráfica',  estado='Activa',     tiene_calidad=True,  estacion_calidad='Paso de La Balsa'),
    'LA BOLSA':                 dict(latitud=3.205007, longitud=-76.494065, tipo='Limnigráfica',  estado='Suspendida', tiene_calidad=True,  estacion_calidad='Paso de La Bolsa'),
    'HORMIGUERO':               dict(latitud=3.303185, longitud=-76.476845, tipo='Limnigráfica',  estado='Suspendida', tiene_calidad=True,  estacion_calidad='Hormiguero'),
    'JUANCHITO (S HASTA 2017)': dict(latitud=3.451685, longitud=-76.475783, tipo='Limnimétrica',  estado='Activa',     tiene_calidad=True,  estacion_calidad='Juanchito'),
    'PUERTO MALLARINO':         dict(latitud=3.444186, longitud=-76.476523, tipo='Limnigráfica',  estado='Activa',     tiene_calidad=False, estacion_calidad=None),
    'PASO LA TORRE':            dict(latitud=3.492342, longitud=-76.483406, tipo='Limnigráfica',  estado='Suspendida', tiene_calidad=True,  estacion_calidad='Paso del Comercio'),
    'GARZONERO SUR (S - 2023)': dict(latitud=4.000335, longitud=-76.312363, tipo='Limnimétrica',  estado='Activa',     tiene_calidad=False, estacion_calidad=None),
    'GARZONERO NORTE':          dict(latitud=4.016306, longitud=-76.311949, tipo='Limnimétrica',  estado='Suspendida', tiene_calidad=False, estacion_calidad=None),
    'MEDICANOA':                dict(latitud=3.891038, longitud=-76.348377, tipo='Limnigráfica',  estado='Activa',     tiene_calidad=True,  estacion_calidad='Mediacanoa'),
    'PUERTO PEDRERO (S - 2017)':dict(latitud=4.318367, longitud=-76.171784, tipo='Limnimétrica',  estado='Suspendida', tiene_calidad=False, estacion_calidad=None),
    'GUAYABAL (HASTA 2022)':    dict(latitud=4.409939, longitud=-76.099568, tipo='Limnimétrica',  estado='Activa',     tiene_calidad=True,  estacion_calidad='Puente Guayabal'),
    'LA VICTORIA':              dict(latitud=4.524927, longitud=-76.043846, tipo='Limnigráfica',  estado='Activa',     tiene_calidad=True,  estacion_calidad='La Victoria'),
    'JUAN DIAZ':                dict(latitud=4.651002, longitud=-76.023158, tipo='Limnimétrica',  estado='Suspendida', tiene_calidad=False, estacion_calidad=None),
    'ANACARO':                  dict(latitud=4.784274, longitud=-75.96926,  tipo='Limnigráfica',  estado='Activa',     tiene_calidad=True,  estacion_calidad='Anacaro'),
}

# Columnas ENE..DIC ocupan columnas 2..13 del Excel (col 1 = DÍA)
MONTH_COLS = list(range(2, 14))   # 12 meses


def _num(s):
    """'1,490.0' → 1490.0 ; '405.0' → 405.0 ; None si no parsea."""
    if s is None:
        return None
    t = str(s).replace(',', '').strip()
    m = re.search(r'-?\d+(?:\.\d+)?', t)
    return float(m.group()) if m else None


def parse_umbrales(path):
    """Extrae estadísticas de umbrales_caudal.txt; campos faltantes → None."""
    out = dict(n_dias=None, promedio_m3s=None, mediana_m3s=None,
               minimo_m3s=None, maximo_m3s=None,
               umbral_invierno_m3s=None, umbral_verano_m3s=None)
    if not os.path.isfile(path):
        return out

    with open(path, encoding='utf-8') as fh:
        txt = fh.read()

    m = re.search(r'Datos:\s*([\d.,]+)\s*caudales', txt)
    if m:
        out['n_dias'] = int(_num(m.group(1)))

    patterns = {
        'promedio_m3s':        r'Promedio\s*:\s*([\d.,]+)',
        'mediana_m3s':         r'Mediana\s*:\s*([\d.,]+)',
        'minimo_m3s':          r'M[ií]nimo\s*:\s*([\d.,]+)',
        'maximo_m3s':          r'M[áa]ximo\s*:\s*([\d.,]+)',
        'umbral_invierno_m3s': r'INVIERNO\s*:\s*caudal\s*[≥>=]+\s*([\d.,]+)',
        'umbral_verano_m3s':   r'VERANO\s*:\s*caudal\s*[≤<=]+\s*([\d.,]+)',
    }
    for key, pat in patterns.items():
        mm = re.search(pat, txt)
        if mm:
            out[key] = _num(mm.group(1))
    return out


def read_caudal(xlsx_path):
    """Lee todas las hojas (año) → lista de {fecha:'YYYY-MM-DD', caudal:float}.
    Devuelve (registros_ordenados, años_con_datos). Solo incluye en años_con_datos
    aquellos años que tienen al menos un valor numérico real (no S/D)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    records = []
    years_with_data = []
    for sheet in wb.sheetnames:
        try:
            year = int(re.search(r'\d{4}', sheet).group())
        except (AttributeError, ValueError):
            continue
        ws = wb[sheet]
        year_records = []
        for r in range(3, 34):                       # filas 3..33 = días 1..31
            day = ws.cell(row=r, column=1).value
            try:
                day = int(day)
            except (TypeError, ValueError):
                continue
            for month, col in enumerate(MONTH_COLS, start=1):
                val = ws.cell(row=r, column=col).value
                if val is None or (isinstance(val, str) and val.strip().upper() in ('S/D', '')):
                    continue
                try:
                    caudal = float(val)
                except (TypeError, ValueError):
                    continue
                try:
                    fecha = datetime.date(year, month, day)   # valida día real
                except ValueError:
                    continue
                year_records.append((fecha, caudal))
        if year_records:                             # solo años con datos reales
            records.extend(year_records)
            years_with_data.append(year)
    wb.close()
    records.sort(key=lambda x: x[0])
    return records, sorted(set(years_with_data))


def main():
    if not os.path.isdir(SRC_DIR):
        raise SystemExit(f'No existe la carpeta fuente:\n  {SRC_DIR}')

    os.makedirs(OUT_DIR, exist_ok=True)

    estaciones = []
    total_dias = 0
    imagenes   = 0

    for folder in sorted(os.listdir(SRC_DIR)):
        src_station = os.path.join(SRC_DIR, folder)
        if not os.path.isdir(src_station):
            continue

        xlsx = os.path.join(src_station, 'caudal_2010.xlsx')
        if not os.path.isfile(xlsx):
            continue   # carpeta sin datos de caudal → ignorar silenciosamente

        if folder in EXCLUIR:
            continue   # excluidas explícitamente (sin datos reales)

        if folder not in META:
            print(f'  ⚠ WARNING: "{folder}" sin entrada en META — omitida')
            continue

        meta = META[folder]

        # Umbrales / estadísticas
        umbrales = parse_umbrales(os.path.join(src_station, 'umbrales_caudal.txt'))

        # CDC
        cdc_src = os.path.join(src_station, 'curva_duracion_caudales.png')
        tiene_cdc = os.path.isfile(cdc_src)

        # Serie de caudal
        records, years = read_caudal(xlsx)
        total_dias += len(records)
        anios = f'{years[0]}–{years[-1]}' if years else None

        # Carpeta de salida por estación
        out_station = os.path.join(OUT_DIR, folder)
        os.makedirs(out_station, exist_ok=True)

        # CSV de caudal diario
        with open(os.path.join(out_station, 'caudal_diario.csv'),
                  'w', newline='', encoding='utf-8') as fh:
            w = csv.writer(fh)
            w.writerow(['FECHA', 'CAUDAL_M3S'])
            for fecha, caudal in records:
                w.writerow([fecha.isoformat(), caudal])

        # Copiar CDC
        if tiene_cdc:
            shutil.copyfile(cdc_src, os.path.join(out_station, 'curva_duracion_caudales.png'))
            imagenes += 1

        estaciones.append({
            'nombre':           folder,
            'nombre_display':   NOMBRE_DISPLAY.get(folder, folder),
            'nombre_cvc':       None,
            'latitud':          meta['latitud'],
            'longitud':         meta['longitud'],
            'tipo':             meta['tipo'],
            'estado':           meta['estado'],
            'tiene_calidad':    meta['tiene_calidad'],
            'estacion_calidad': meta['estacion_calidad'],
            'tiene_cdc':        tiene_cdc,
            'años_datos':       anios,
            **umbrales,
        })
        print(f'  ✓ {folder:<26} {len(records):>5} días · {anios or "—"}'
              f'{" · CDC" if tiene_cdc else ""}')

    with open(os.path.join(OUT_DIR, 'estaciones_hidro.json'),
              'w', encoding='utf-8') as fh:
        json.dump(estaciones, fh, ensure_ascii=False, indent=2)

    print('\n── Resumen ─────────────────────────────────')
    print(f'  Estaciones procesadas : {len(estaciones)}')
    print(f'  Días de caudal totales: {total_dias:,}')
    print(f'  Imágenes CDC copiadas : {imagenes}')
    print(f'  Salida                : {OUT_DIR}')


if __name__ == '__main__':
    main()
