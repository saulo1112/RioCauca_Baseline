# -*- coding: utf-8 -*-
"""build_calidad_trib.py — Cruce de calidad del agua × puntos de monitoreo (tributarios).

Se ejecuta UNA SOLA VEZ desde la terminal para generar los archivos estáticos
que el visor web consume:

    py -3.13 src/build_calidad_trib.py

Cruza las muestras fisicoquímicas del Excel consolidado con los puntos
georreferenciados (WGS84) del GeoJSON, emparejando por el nombre del punto de
monitoreo normalizado (el campo Rio del Excel no es fiable para el join porque
etiqueta subcorrientes del Risaralda por su propio nombre). Produce:

    data/geovisor/puntos_calidad_tributarios.geojson   (capa del mapa, 76 puntos)
    data/geovisor/estadisticas_puntos.csv              (resumen media/min/max/p90)
    data/geovisor/csv_por_punto/<Rio>_<Punto>.csv      (una serie por estación)

El GeoJSON incrusta en cada feature los valores más recientes de 10 parámetros
prioritarios (como string JSON en `parametros_recientes`) y el nombre del CSV
de descarga (`csv_filename`), de modo que el popup del visor no necesita parsear
el Excel en el navegador.
"""

import os
import re
import sys
import csv
import json
import unicodedata
import datetime

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, ValueError):
    pass

# ── Rutas ──────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
XLSX_PATH = os.path.join(PROJECT_DIR, 'data', 'databases', 'Calidad_agua_completo_v12.xlsx')
GEO_PATH  = os.path.join(PROJECT_DIR, 'data', 'databases', 'Calidad_tributarios.geojson')
OUT_DIR   = os.path.join(PROJECT_DIR, 'data', 'geovisor')
CSV_DIR   = os.path.join(OUT_DIR, 'csv_por_punto')
GEOJSON_OUT = os.path.join(OUT_DIR, 'puntos_calidad_tributarios.geojson')
STATS_OUT   = os.path.join(OUT_DIR, 'estadisticas_puntos.csv')

SHEET       = 'CONSOLIDADO'
HEADER_ROW  = 3            # el encabezado real está en la fila 3
FIRST_DATA  = 4
COL_RIO     = 'Rio'
COL_PUNTO   = 'Punto de Monitoreo'
COL_ANIO    = 'Año'
COL_FECHA   = 'Fecha Muestreo'

# 10 parámetros prioritarios (nombres exactos del encabezado) → clave corta
PRIORITY = [
    'pH Campo',
    'Temperatura Agua (°C)',
    'Turbiedad (UNT)',
    'Sólidos Suspendidos Totales (mg/l)',
    'DBO (mg O2/l)',
    'DQO (mg O2/l)',
    'Oxígeno Disuelto (mg O2/l)',
    'Nitrógeno Total (mg N/l)',
    'Fósforo Total (mg P/l)',
    'Coliformes Fecales (NMP/100ml)',
]


# ── Helpers ────────────────────────────────────────────────────────────
def norm_punto(s):
    """Clave de cruce: minúsculas, sin acentos, separadores , - – → espacio."""
    if s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s).strip()).encode('ascii', 'ignore').decode().lower()
    s = re.sub(r'[,\-–]+', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()


def to_num(v):
    """Valor numérico para estadísticas; S/D/vacío → None; '<X'/'>X' → float(X)."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip()
    if t == '' or t.upper() in ('S/D', 'S/D.', 'N.D.', 'N.D', '*'):
        return None
    if t and t[0] in '<>':
        t = t[1:].strip()
    t = t.replace(',', '')
    m = re.search(r'-?\d+(?:\.\d+)?', t)
    return float(m.group()) if m else None


def norm_filename(rio, punto):
    """'{Rio}_{Punto}.csv' ascii-safe, sin dobles '_', nombre completo máx 60 chars."""
    base = f'{rio}_{punto}'
    base = unicodedata.normalize('NFKD', base).encode('ascii', 'ignore').decode()
    base = re.sub(r'[^A-Za-z0-9]+', '_', base).strip('_')
    base = re.sub(r'_+', '_', base)[:56].strip('_')   # 56 + '.csv' = 60
    return base + '.csv'


def parse_fecha(v):
    """Devuelve datetime.date para ordenar; formatos comunes o None."""
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date):
        return v
    s = str(v or '').strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y'):
        try:
            return datetime.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return datetime.date(1900, 1, 1)


def cell_str(v):
    """Valor original para el CSV por punto (conserva S/D, <X)."""
    if v is None:
        return ''
    if isinstance(v, datetime.datetime):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, datetime.date):
        return v.strftime('%d/%m/%Y')
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def p90(values):
    """Percentil 90 (interpolación lineal) de una lista ya filtrada de floats."""
    if not values:
        return None
    xs = sorted(values)
    if len(xs) == 1:
        return xs[0]
    rank = 0.90 * (len(xs) - 1)
    lo = int(rank)
    frac = rank - lo
    if lo + 1 < len(xs):
        return xs[lo] + frac * (xs[lo + 1] - xs[lo])
    return xs[lo]


# ── Carga ──────────────────────────────────────────────────────────────
def load_geojson_points():
    with open(GEO_PATH, encoding='utf-8') as fh:
        gj = json.load(fh)
    pts = {}
    for feat in gj['features']:
        p = feat['properties']
        pts[norm_punto(p.get('Punto_Monitoreo'))] = p
    return pts


def load_excel():
    """Devuelve (headers, rows_por_punto). rows_por_punto: dict[norm] → list[row]."""
    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)
    ws = wb[SHEET]
    header = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[0]
    headers = [h if h is not None else '' for h in header]
    idx_punto = headers.index(COL_PUNTO)

    groups = {}
    for row in ws.iter_rows(min_row=FIRST_DATA, values_only=True):
        if row[0] is None and row[idx_punto] is None:
            continue
        key = norm_punto(row[idx_punto])
        groups.setdefault(key, []).append(list(row))
    wb.close()
    return headers, groups


# ── Main ───────────────────────────────────────────────────────────────
def main():
    if not os.path.isfile(XLSX_PATH):
        raise SystemExit(f'No existe el Excel:\n  {XLSX_PATH}')
    if not os.path.isfile(GEO_PATH):
        raise SystemExit(f'No existe el GeoJSON:\n  {GEO_PATH}')

    os.makedirs(CSV_DIR, exist_ok=True)

    points = load_geojson_points()
    headers, groups = load_excel()
    print(f'GeoJSON: {len(points)} puntos | Excel: {sum(len(v) for v in groups.values())} muestras '
          f'en {len(groups)} puntos')

    idx = {name: headers.index(name) for name in
           (COL_RIO, COL_PUNTO, COL_ANIO, COL_FECHA) if name in headers}
    prio_idx = {name: headers.index(name) for name in PRIORITY if name in headers}
    faltan = [name for name in PRIORITY if name not in prio_idx]
    if faltan:
        print('  ⚠ Parámetros no encontrados en el encabezado:', faltan)

    features = []
    stats_rows = []
    total_muestras = 0
    sin_cruce = []

    for key, p in points.items():
        rio   = p.get('Rio')
        punto = p.get('Punto_Monitoreo')
        rows  = groups.get(key, [])
        if not rows:
            sin_cruce.append(punto)

        rows.sort(key=lambda r: parse_fecha(r[idx[COL_FECHA]]))
        total_muestras += len(rows)

        csv_filename = norm_filename(rio, punto)

        # 1. CSV por punto (todas las columnas, valores originales)
        with open(os.path.join(CSV_DIR, csv_filename), 'w', newline='', encoding='utf-8-sig') as fh:
            w = csv.writer(fh)
            w.writerow(headers)
            for r in rows:
                w.writerow([cell_str(v) for v in r])

        # 2. Años y valores recientes
        anios = [int(r[idx[COL_ANIO]]) for r in rows
                 if isinstance(r[idx[COL_ANIO]], (int, float))]
        anio_min = min(anios) if anios else None
        anio_max = max(anios) if anios else None

        recientes = []
        for name, i in prio_idx.items():
            valor, anio = None, None
            for r in reversed(rows):          # rows están ascendentes por fecha
                if to_num(r[i]) is not None:
                    valor = cell_str(r[i])
                    anio = int(r[idx[COL_ANIO]]) if isinstance(r[idx[COL_ANIO]], (int, float)) else None
                    break
            if valor is not None:
                recientes.append({'param': name, 'valor': valor, 'anio': anio})

        # 3. Feature GeoJSON
        features.append({
            'type': 'Feature',
            'geometry': {'type': 'Point',
                         'coordinates': [p.get('Longitud_WGS84'), p.get('Latitud_WGS84')]},
            'properties': {
                'Rio':                 rio,
                'Punto_Monitoreo':     punto,
                'Latitud_WGS84':       p.get('Latitud_WGS84'),
                'Longitud_WGS84':      p.get('Longitud_WGS84'),
                'N_Registros':         len(rows),
                'Año_Min':             anio_min,
                'Año_Max':             anio_max,
                'Fuente_CRS':          p.get('Fuente_CRS'),
                'csv_filename':        csv_filename,
                'parametros_recientes': json.dumps(recientes, ensure_ascii=False),
            },
        })

        # 4. Fila de estadísticas
        srow = {'Rio': rio, 'Punto_Monitoreo': punto, 'N_Registros': len(rows),
                'Año_Min': anio_min, 'Año_Max': anio_max}
        for name, i in prio_idx.items():
            vals = [to_num(r[i]) for r in rows]
            vals = [v for v in vals if v is not None]
            if vals:
                srow[f'{name} | media'] = round(sum(vals) / len(vals), 4)
                srow[f'{name} | min']   = round(min(vals), 4)
                srow[f'{name} | max']   = round(max(vals), 4)
                srow[f'{name} | p90']   = round(p90(vals), 4)
            else:
                srow[f'{name} | media'] = srow[f'{name} | min'] = None
                srow[f'{name} | max']   = srow[f'{name} | p90'] = None
        stats_rows.append(srow)

    # ── Escritura de salidas ────────────────────────────────────────────
    geojson_out = {'type': 'FeatureCollection',
                   'crs': {'type': 'name', 'properties': {'name': 'urn:ogc:def:crs:OGC:1.3:CRS84'}},
                   'features': features}
    with open(GEOJSON_OUT, 'w', encoding='utf-8') as fh:
        json.dump(geojson_out, fh, ensure_ascii=False, separators=(',', ':'))

    stat_cols = ['Rio', 'Punto_Monitoreo', 'N_Registros', 'Año_Min', 'Año_Max']
    for name in prio_idx:
        stat_cols += [f'{name} | media', f'{name} | min', f'{name} | max', f'{name} | p90']
    with open(STATS_OUT, 'w', newline='', encoding='utf-8-sig') as fh:
        w = csv.DictWriter(fh, fieldnames=stat_cols)
        w.writeheader()
        for srow in stats_rows:
            w.writerow(srow)

    # ── Resumen ─────────────────────────────────────────────────────────
    print('\n── Resumen ─────────────────────────────────')
    print(f'  Puntos procesados        : {len(features)}')
    print(f'  CSVs por punto           : {len(features)}')
    print(f'  Muestras distribuidas    : {total_muestras}')
    print(f'  Puntos sin cruce (0 obs) : {len(sin_cruce)}'
          + (f' → {sin_cruce}' if sin_cruce else ''))
    print(f'  GeoJSON                  : {GEOJSON_OUT}')
    print(f'  Estadísticas             : {STATS_OUT}')
    print(f'  Carpeta CSVs             : {CSV_DIR}')


if __name__ == '__main__':
    main()
