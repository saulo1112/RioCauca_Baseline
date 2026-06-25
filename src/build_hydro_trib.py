# -*- coding: utf-8 -*-
"""build_hydro_trib.py — Preparación de datos de estaciones hidrométricas de ríos tributarios.

Se ejecuta UNA SOLA VEZ desde la terminal para generar los archivos estáticos
que el visor web consume:

    python src/build_hydro_trib.py

Lee la carpeta fuente (Excel de caudal diario + umbrales + CDC por estación),
la cruza con Estaciones_tributarios.geojson para obtener coordenadas, y produce:

    data/hydrology/estaciones_hidro_trib.json       (metadatos + estadísticas)
    data/hydrology/tributarios/<NOMBRE>/caudal_diario.csv
    data/hydrology/tributarios/<NOMBRE>/curva_duracion_caudales.png
"""

import os
import re
import sys
import csv
import json
import shutil
import datetime

import openpyxl

try:
    sys.stdout.reconfigure(encoding='utf-8')
except (AttributeError, ValueError):
    pass

# ── Rutas ──────────────────────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)
SRC_DIR = os.path.join(
    PROJECT_DIR, 'data', 'databases', 'hidrology',
    'Estaciones Hidroclimatológicas - Ríos tributarios',
)
GEO_PATH = os.path.join(
    PROJECT_DIR, 'data', 'databases', 'Estaciones_tributarios.geojson',
)
OUT_DIR = os.path.join(PROJECT_DIR, 'data', 'hydrology', 'tributarios')
JSON_OUT = os.path.join(PROJECT_DIR, 'data', 'hydrology', 'estaciones_hidro_trib.json')

# Carpetas de primer nivel a ignorar (no son ríos)
IGNORAR_NIVEL1 = {'Curvas de duración de caudal', 'ZABALETAS (sin info)', 'RIO RISARALDA'}

# Columnas ENE..DIC ocupan columnas 2..13 (col 1 = DÍA) — mismo formato que Cauca
MONTH_COLS = list(range(2, 14))

# Nombre de mes → número (para leer el Excel de Risaralda)
MESES_RDA = {'ENE':1,'FEB':2,'MAR':3,'ABR':4,'MAY':5,'JUN':6,
             'JUL':7,'AGO':8,'SEP':9,'OCT':10,'NOV':11,'DIC':12}

# Estaciones del Río Risaralda con rutas absolutas en SRC_DIR/RIO RISARALDA/
RISARALDA_STATIONS = [
    {
        'geo_key':       'CASA MAQUINAS',
        'caudal_xlsx':   'CAUDAL/Casa_Maquinas_Rda_caudal.xlsx',
        'umbrales_txt':  'umbrales_caudal_Casa_Maquinas.txt',
        'cdc_png':       'Gráficas/curva_duracion_Casa_Maquinas.png',
    },
    {
        'geo_key':       'RIO RISARALDA EHT',
        'caudal_xlsx':   'CAUDAL/Rio_Risaralda_caudal.xlsx',
        'umbrales_txt':  'umbrales_caudal_Rio_Risaralda_EHT.txt',
        'cdc_png':       'Gráficas/curva_duracion_Rio_Risaralda_EHT.png',
    },
]


def _num(s):
    """'1,490.0' → 1490.0 ; None si no parsea."""
    if s is None:
        return None
    t = str(s).replace(',', '').strip()
    m = re.search(r'-?\d+(?:\.\d+)?', t)
    return float(m.group()) if m else None


def _norm_folder(name):
    """'AMAIME (Hasta 2023)' → 'AMAIME'  — elimina sufijos entre paréntesis."""
    return re.sub(r'\s*\([^)]*\)', '', name).strip().upper()


def parse_umbrales(path):
    """Extrae estadísticas de umbrales_caudal.txt; campos faltantes → None.
    Soporta tanto 'Datos: X caudales' (formato Cauca) como
    'Registros: X caudales diarios' (formato tributarios)."""
    out = dict(n_dias=None, promedio_m3s=None, mediana_m3s=None,
               minimo_m3s=None, maximo_m3s=None,
               umbral_invierno_m3s=None, umbral_verano_m3s=None)
    if not os.path.isfile(path):
        return out

    with open(path, encoding='utf-8') as fh:
        txt = fh.read()

    # 'Datos: 5868 caudales' o 'Registros: 2,906 caudales diarios'
    m = re.search(r'(?:Datos|Registros)\s*:\s*([\d.,]+)\s*caudales', txt)
    if m:
        out['n_dias'] = int(_num(m.group(1)))

    patterns = {
        'promedio_m3s':        r'Promedio\s*:\s*([\d.,]+)',
        'mediana_m3s':         r'Mediana\s*:\s*([\d.,]+)',
        'minimo_m3s':          r'M[ií]nimo\s*:\s*([\d.,]+)',
        'maximo_m3s':          r'M[áa]ximo\s*:\s*([\d.,]+)',
        'umbral_invierno_m3s': r'INVIERNO\s*:\s*caudal\s*[≥>=]+\s*([\d.,]+)',
        'umbral_verano_m3s':   r'VERANO\s*:\s*caudal\s*[≤<=]+\s*([\d.,]+)',
    }
    for key, pat in patterns.items():
        mm = re.search(pat, txt)
        if mm:
            out[key] = _num(mm.group(1))
    return out


def read_caudal(xlsx_path):
    """Lee todas las hojas (año) → lista de (fecha, caudal).
    Igual que en build_hydro_data.py: hojas por año, 12 columnas de meses, 31 filas."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    records = []
    years_with_data = []
    for sheet in wb.sheetnames:
        try:
            year = int(re.search(r'\d{4}', sheet).group())
        except (AttributeError, ValueError):
            continue
        ws = wb[sheet]
        year_records = []
        for r in range(3, 34):
            day = ws.cell(row=r, column=1).value
            try:
                day = int(day)
            except (TypeError, ValueError):
                continue
            for month, col in enumerate(MONTH_COLS, start=1):
                val = ws.cell(row=r, column=col).value
                if val is None or (isinstance(val, str) and val.strip().upper() in ('S/D', '')):
                    continue
                try:
                    caudal = float(val)
                except (TypeError, ValueError):
                    continue
                try:
                    fecha = datetime.date(year, month, day)
                except ValueError:
                    continue
                year_records.append((fecha, caudal))
        if year_records:
            records.extend(year_records)
            years_with_data.append(year)
    wb.close()
    records.sort(key=lambda x: x[0])
    return records, sorted(set(years_with_data))


def read_caudal_risaralda(xlsx_path):
    """Lee el Excel de Risaralda (formato CMáx/CMed/CMín por mes) → (registros, años).
    Extrae únicamente CMed (caudal promedio diario)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    records = []
    years_with_data = []

    for sheet in wb.sheetnames:
        try:
            year = int(re.search(r'\d{4}', sheet).group())
        except (AttributeError, ValueError):
            continue
        ws = wb[sheet]
        all_rows = list(ws.iter_rows(values_only=True))
        if len(all_rows) < 4:
            continue

        # Fila 2 (índice 1): nombres de mes en las posiciones de primera subcolumna
        month_row = all_rows[1]
        # Mapear mes → índice de CMed (0-based): CMed está 1 posición tras el encabezado del mes
        month_cmed = {}
        for i, val in enumerate(month_row):
            if val and str(val).strip().upper() in MESES_RDA:
                month_num = MESES_RDA[str(val).strip().upper()]
                month_cmed[month_num] = i + 1   # CMed = columna siguiente al encabezado

        year_records = []
        for row in all_rows[3:]:   # datos desde fila índice 3 (cuarta fila de la hoja)
            if not row or row[0] is None:
                continue
            try:
                day = int(row[0])
            except (TypeError, ValueError):
                continue
            for month, cmed_idx in month_cmed.items():
                if cmed_idx >= len(row):
                    continue
                val = row[cmed_idx]
                if val is None:
                    continue
                try:
                    caudal = float(val)
                except (TypeError, ValueError):
                    continue
                try:
                    fecha = datetime.date(year, month, day)
                except ValueError:
                    continue
                year_records.append((fecha, caudal))

        if year_records:
            records.extend(year_records)
            years_with_data.append(year)

    wb.close()
    records.sort(key=lambda x: x[0])
    return records, sorted(set(years_with_data))


def parse_umbrales_risaralda(path):
    """Extrae umbrales del formato Risaralda:
       'Umbral Invierno (Q30%): X m3/s' / 'Umbral Verano (Q70%): X m3/s'."""
    out = dict(n_dias=None, umbral_invierno_m3s=None, umbral_verano_m3s=None)
    if not os.path.isfile(path):
        return out
    with open(path, encoding='utf-8') as fh:
        txt = fh.read()
    m = re.search(r'Datos:\s*([\d.,]+)\s*valores', txt)
    if m:
        out['n_dias'] = int(_num(m.group(1)))
    m = re.search(r'Umbral Invierno[^:]*:\s*([\d.,]+)', txt)
    if m:
        out['umbral_invierno_m3s'] = _num(m.group(1))
    m = re.search(r'Umbral Verano[^:]*:\s*([\d.,]+)', txt)
    if m:
        out['umbral_verano_m3s'] = _num(m.group(1))
    return out


def compute_stats(records):
    """Calcula promedio, mediana, mínimo y máximo a partir de los registros diarios."""
    if not records:
        return dict(promedio_m3s=None, mediana_m3s=None, minimo_m3s=None, maximo_m3s=None)
    values = sorted(c for _, c in records)
    n = len(values)
    median = values[n // 2] if n % 2 else (values[n // 2 - 1] + values[n // 2]) / 2
    return dict(
        promedio_m3s=round(sum(values) / n, 2),
        mediana_m3s=round(median, 2),
        minimo_m3s=round(values[0], 2),
        maximo_m3s=round(values[-1], 2),
    )


def load_geojson():
    """Carga el GeoJSON y devuelve un dict: ESTACION_UPPER → propiedades."""
    with open(GEO_PATH, encoding='utf-8') as fh:
        gj = json.load(fh)
    geo = {}
    for feat in gj['features']:
        p = feat['properties']
        key = p['Estacion'].strip().upper()
        geo[key] = {
            'estacion':   p['Estacion'],
            'rio':        p['Rio'],
            'municipio':  p.get('Municipio') or '',
            'latitud':    p['Latitud'],
            'longitud':   p['Longitud'],
            'suspendida': p.get('Suspendida', 'No') == 'Si',
        }
    return geo


def main():
    if not os.path.isdir(SRC_DIR):
        raise SystemExit(f'No existe la carpeta fuente:\n  {SRC_DIR}')
    if not os.path.isfile(GEO_PATH):
        raise SystemExit(f'No existe el GeoJSON:\n  {GEO_PATH}')

    os.makedirs(OUT_DIR, exist_ok=True)

    geo = load_geojson()
    print(f'GeoJSON: {len(geo)} estaciones cargadas')

    # Inicializar todas las estaciones del GeoJSON con datos básicos
    # (serán actualizadas si se encuentran datos de caudal)
    estaciones = {}
    for key, g in geo.items():
        nombre = g['estacion']
        estaciones[key] = {
            'nombre':               nombre,
            'nombre_display':       nombre,
            'rio':                  g['rio'],
            'municipio':            g['municipio'],
            'latitud':              g['latitud'],
            'longitud':             g['longitud'],
            'estado':               'Suspendida' if g['suspendida'] else 'Activa',
            'ruta_hidro':           f'tributarios/{nombre}',
            'tiene_cdc':            False,
            'años_datos':           None,
            'n_dias':               None,
            'promedio_m3s':         None,
            'mediana_m3s':          None,
            'minimo_m3s':           None,
            'maximo_m3s':           None,
            'umbral_invierno_m3s':  None,
            'umbral_verano_m3s':    None,
        }

    total_dias = 0
    imagenes   = 0
    procesadas = 0

    # Recorrer carpetas de río → estación
    for rio_folder in sorted(os.listdir(SRC_DIR)):
        if rio_folder in IGNORAR_NIVEL1:
            continue
        rio_path = os.path.join(SRC_DIR, rio_folder)
        if not os.path.isdir(rio_path):
            continue

        for est_folder in sorted(os.listdir(rio_path)):
            est_path = os.path.join(rio_path, est_folder)
            if not os.path.isdir(est_path):
                continue

            xlsx = os.path.join(est_path, 'caudal.xlsx')
            if not os.path.isfile(xlsx):
                continue   # sin datos de caudal → ignorar

            # Hacer match con GeoJSON por nombre normalizado
            norm = _norm_folder(est_folder)
            if norm not in estaciones:
                print(f'  ⚠ WARNING: "{est_folder}" (norm: "{norm}") sin entrada en GeoJSON — omitida')
                continue

            g = estaciones[norm]
            nombre = g['nombre']

            # Umbrales / estadísticas
            umbrales = parse_umbrales(os.path.join(est_path, 'umbrales_caudal.txt'))

            # CDC
            cdc_src = os.path.join(est_path, 'curva_duracion.png')
            tiene_cdc = os.path.isfile(cdc_src)

            # Serie de caudal
            try:
                records, years = read_caudal(xlsx)
            except Exception as exc:
                print(f'  ✗ ERROR leyendo {xlsx}: {exc}')
                continue

            total_dias += len(records)
            anios = f'{years[0]}–{years[-1]}' if years else None

            # Carpeta de salida
            out_station = os.path.join(OUT_DIR, nombre)
            os.makedirs(out_station, exist_ok=True)

            # CSV de caudal diario
            with open(os.path.join(out_station, 'caudal_diario.csv'),
                      'w', newline='', encoding='utf-8') as fh:
                w = csv.writer(fh)
                w.writerow(['FECHA', 'CAUDAL_M3S'])
                for fecha, caudal in records:
                    w.writerow([fecha.isoformat(), caudal])

            # Copiar CDC
            if tiene_cdc:
                shutil.copyfile(cdc_src, os.path.join(out_station, 'curva_duracion_caudales.png'))
                imagenes += 1

            # Actualizar entrada en el dict
            estaciones[norm].update({
                'tiene_cdc':           tiene_cdc,
                'años_datos':          anios,
                **umbrales,
            })
            procesadas += 1
            print(f'  ✓ {nombre:<22} [{rio_folder}] {len(records):>5} días · {anios or "—"}'
                  f'{" · CDC" if tiene_cdc else ""}')

    # ── Estaciones especiales: Río Risaralda ────────────────────────────
    rda_base = os.path.join(SRC_DIR, 'RIO RISARALDA')
    for rda in RISARALDA_STATIONS:
        key = rda['geo_key']
        if key not in estaciones:
            print(f'  ⚠ WARNING: Risaralda "{key}" sin entrada en GeoJSON — omitida')
            continue

        nombre = estaciones[key]['nombre']
        xlsx   = os.path.join(rda_base, rda['caudal_xlsx'])
        txt    = os.path.join(rda_base, rda['umbrales_txt'])
        cdc_src = os.path.join(rda_base, rda['cdc_png'])

        if not os.path.isfile(xlsx):
            print(f'  ⚠ WARNING: {xlsx} no encontrado — omitido')
            continue

        try:
            records, years = read_caudal_risaralda(xlsx)
        except Exception as exc:
            print(f'  ✗ ERROR leyendo {xlsx}: {exc}')
            continue

        umbrales  = parse_umbrales_risaralda(txt)
        stats     = compute_stats(records)
        tiene_cdc = os.path.isfile(cdc_src)
        anios     = f'{years[0]}–{years[-1]}' if years else None

        total_dias += len(records)

        out_station = os.path.join(OUT_DIR, nombre)
        os.makedirs(out_station, exist_ok=True)

        with open(os.path.join(out_station, 'caudal_diario.csv'),
                  'w', newline='', encoding='utf-8') as fh:
            w = csv.writer(fh)
            w.writerow(['FECHA', 'CAUDAL_M3S'])
            for fecha, caudal in records:
                w.writerow([fecha.isoformat(), caudal])

        if tiene_cdc:
            shutil.copyfile(cdc_src, os.path.join(out_station, 'curva_duracion_caudales.png'))
            imagenes += 1

        # n_dias del txt tiene prioridad; si no, usar longitud real del CSV
        n_dias_final = umbrales.pop('n_dias') or len(records)
        estaciones[key].update({
            'tiene_cdc':   tiene_cdc,
            'años_datos':  anios,
            'n_dias':      n_dias_final,
            **stats,
            **umbrales,
        })
        procesadas += 1
        print(f'  ✓ {nombre:<22} [RIO RISARALDA] {len(records):>5} días · {anios or "—"}'
              f'{" · CDC" if tiene_cdc else ""}')

    # Guardar JSON (mismo orden que el GeoJSON)
    output = list(estaciones.values())
    with open(JSON_OUT, 'w', encoding='utf-8') as fh:
        json.dump(output, fh, ensure_ascii=False, indent=2)

    print('\n── Resumen ─────────────────────────────────')
    print(f'  Estaciones en GeoJSON    : {len(geo)}')
    print(f'  Estaciones con caudal    : {procesadas}')
    print(f'  Días de caudal totales   : {total_dias:,}')
    print(f'  Imágenes CDC copiadas    : {imagenes}')
    print(f'  JSON de salida           : {JSON_OUT}')
    print(f'  Carpeta de salida        : {OUT_DIR}')


if __name__ == '__main__':
    main()
